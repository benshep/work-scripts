import asyncio
import os
import sys
import pandas
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from rayleigh_connect import EasyAPI
from rayleigh_api_key import api_key
from work_folders import docs_folder, misc_folder
from work_tools import read_excel

sys.path.append(os.path.join(misc_folder, 'Scripts'))
from get_energy_usage import get_co2_data

rayleigh = EasyAPI(api_key)


async def find_sensors(names: list[str]) -> dict[str, tuple[str, str]]:
    """Returns the gateway and sensor IDs for the given sensor names."""
    return_dict = {}
    for gateway in await rayleigh.fetch_gateways():
        for sensor in await rayleigh.fetch_sensors(gateway['id']):
            if sensor['name'] in names:
                return_dict[sensor['name']] = (gateway['id'], sensor['id'])
    return return_dict


async def list_sensors():
    """Lists sensor names defined for each gateway."""
    for gateway in await rayleigh.fetch_gateways():
        for sensor in await rayleigh.fetch_sensors(gateway['id']):
            if sensor['name'] not in (None, 'Frequency'):  # lots of sensors called Frequency
                print(gateway['id'], sensor['id'], sensor['name'], sep='\t')


def update_energy_data():
    """Sync wrapper for async archive_data function."""
    asyncio.run(archive_data())


async def archive_data():
    """Fetch new energy and carbon data and drop it into the spreadsheet."""
    excel_filename = os.path.join(docs_folder, 'CLARA', 'CLARA electricity usage.xlsx')
    meters_sheet = 'Meters'
    # Get the list of sensors from the top 3 rows of this Excel file
    workbook = read_excel(excel_filename, header=[0, 1, 2], sheet_name=None, index_col=0)
    meter_data = workbook[meters_sheet]
    # Use the first column of the file to get a start date for our data
    last_timestamp = max(meter_data.index)
    start = last_timestamp - pandas.to_timedelta(1, unit='hours')  # go back a bit, to avoid gaps
    end = last_timestamp + pandas.to_timedelta(14, 'days')
    # print(meter_data.columns[:3])
    frames = await asyncio.gather(
        *[rayleigh.fetch_data(
            gateway,
            sensor + '.kwh',  # want the energy data
            start.to_pydatetime(),
            end.to_pydatetime(),
        ) for gateway, sensor, sensor_name in meter_data.columns])
    # print(frames)
    # Resample so that data points are on half-hour intervals
    frames = [resample(frame, key) for frame, key in zip(frames, meter_data.columns)]
    # What is the last data point common to all datasets? Trying to avoid 'orphans'
    first_new_time = last_timestamp + pandas.to_timedelta(30, unit='minutes')
    last_new_time = min(max(frame.index) for frame in frames if len(frame) > 0)
    # Concatenate all the data together horizontally...
    new_data = pandas.concat(frames, axis=1)
    # Delete all rows with data after this last common point, and remove any overlap
    new_data = new_data.loc[first_new_time:last_new_time]
    print('New data from', first_new_time, 'to', last_new_time)
    # and add it to the pre-existing Excel data
    meter_data = pandas.concat([meter_data, new_data])
    # print(meter_data)

    # Get the carbon intensity data
    intensity_sheet = 'Intensity'
    intensity_data = workbook[intensity_sheet]
    first_new_int_time = max(intensity_data.index) + pandas.to_timedelta(30, 'min')
    new_intensity_data = get_co2_data(first_new_int_time, 'WA4', do_pivot=False, end=last_new_time)
    if not new_intensity_data.empty:
        print('New intensity data from', min(new_intensity_data.index), 'to', max(new_intensity_data.index))
        new_intensity_data.index = new_intensity_data.index.tz_localize(None)  # make timezone-unaware for Excel
        # print(new_intensity_data)
        new_intensity_data.columns = intensity_data.columns
        intensity_data = pandas.concat([intensity_data, new_intensity_data])

    # Write the old and new data back into the Excel workbook
    with pandas.ExcelWriter(excel_filename, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        meter_data.to_excel(writer, sheet_name=meters_sheet, merge_cells=False)
        intensity_data.to_excel(writer, sheet_name=intensity_sheet, merge_cells=True)
    print('Data written to', excel_filename)

    # Carry out some formatting
    workbook = load_workbook(excel_filename)
    bold_format = Font(bold=True)
    word_wrap = Alignment(wrap_text=True)
    for sheet_name in (meters_sheet, intensity_sheet):
        worksheet = workbook[sheet_name]
        worksheet.column_dimensions['A'].width = 20  # otherwise can't read dates
        for col in range(2, worksheet.max_column + 1):
            column_name = worksheet.cell(4, col).column_letter  # use row 4 so it's not merged
            worksheet.column_dimensions[column_name].width = 18
        for row in worksheet.iter_rows(1, 3):
            for cell in row:
                cell.font = bold_format
                if cell.row == 3:
                    cell.alignment = word_wrap
    workbook.save(excel_filename)


def resample(data: list[list], key: tuple[str, str, str]):
    """Convert data in 2d list format to a DataFrame and resample to half-hourly intervals."""
    df = pandas.DataFrame(data, columns=['timestamp', 'reading'])
    df['timestamp'] = pandas.to_datetime(df['timestamp'])
    df.set_index('timestamp', inplace=True)
    df.index = df.index.tz_localize(None)  # make timezone-unaware for Excel
    df = df.resample(pandas.Timedelta('30min')).ffill()
    df = df.dropna()
    df.columns = pandas.MultiIndex.from_tuples([key])
    # print(df.columns)
    # print(df)
    return df


async def test_fetch():
    start = pandas.to_datetime('2025-11-01').to_pydatetime()
    print(start)
    return await rayleigh.fetch_data(
        'Q3025301880080025',
        'e3.kwh',  # want the energy data
        start,
    )

if __name__ == '__main__':
    # print(find_sensors(['PD2 - SW4E - PD934 CLARA Rack Room 2 - Magnet Rack 1',
    #                     'PD2 - SW4D - PD935 CLARA Rack Room 2 - Magnet Rack 2']))
    asyncio.run(list_sensors())

    # data = asyncio.run(test_fetch())
    # print(data)

    # update_energy_data()
